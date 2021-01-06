import openpyxl
from openpyxl.styles import Font
import sys

wb=openpyxl.Workbook()
fontObject = Font(name='Arial',bold=True)
sheet=wb.active
n=int(sys.argv[1])

length=len(sys.argv)
if length > 2:
    print("Not valid")
    sys.quit()

for i in range(1,n+1):
    sheet['A'+str(i+1)] = i
    sheet.cell(row=1 ,column=i+1).value = i

    sheet['A'+str(i+1)].font = fontObject
    sheet.cell(row=1 , column=i+1).font = fontObject

for i in range (1,n+1):
    for j in range (1,n+1):
        sheet.cell(row=j+1 ,column =i+1).value = i*j

wb.save('table.xlsx')


